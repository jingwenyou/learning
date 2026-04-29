import os
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.append(os.path.split(os.path.abspath(os.path.dirname(__file__)))[0])
import time

from common.exception_utils import exception_utils
from common.text_util import read_txt_handel

base_dir = Path(__file__).parent.parent.parent


@exception_utils
class ExcelUtil(object):

    def __init__(self, excel_path='%s/data/case_excel/接口测试用例.xlsx' % base_dir):
        self.wb = load_workbook(excel_path)
        self.base_dir = base_dir

    @exception_utils
    def read_excel(self):
        """读取excel，处理数据，并返回一个格式处理后的字典"""
        value = []
        smoke_value = []
        for sheetname in self.wb.sheetnames:
            ws = self.wb[sheetname]
            case_list = list(ws.values)
            cases_num = len(case_list) - 1  # 一个sheet中用例的数量
            key = case_list.pop(0)  # 提取表头字段作为字典key

            cases_template_list = []
            for i in range(cases_num):  # i：第i个用例
                # casei=list(case_list[i])
                # for j in casei:
                #     if isinstance(j,str) and re.match(r'\{.*\}',j):
                #         try:
                #             casei[j] = json.loads(j)
                #         except Exception as e:
                #             print(e)
                cases_template_list.append(dict(zip(key, case_list[i])))
            value.append({"cases": cases_template_list})

        for v in value:
            for case in v['cases']:
                # print(case)
                if '正常' in str(case):
                    print(case['id'], case)
                    smoke_value.append(case)

        smoke = {"cases": smoke_value}
        return value, smoke

    @exception_utils
    def write_excel(self):
        """运行结果写入excel"""
        l_reponse, l_ispass = read_txt_handel()
        i = 0
        j = 0
        for sheetname in self.wb.sheetnames:
            ws = self.wb[sheetname]
            # 实际结果列
            for col in range(len(ws[2])):
                if ws[1][col].value == 'actual':
                    actual_col = col
                    continue
                if ws[1][col].value == 'validate':
                    validate_col = col
            for row in range(2, ws.max_row + 1):
                ws[row][actual_col].value = l_reponse[i]
                i += 1
                ws[row][validate_col].value = l_ispass[j]
                j += 1

        strftime = time.strftime("%Y%m%d_%H_%M_%S")
        save_path = "%s/output/run_result_excel/运行结果_%s.xlsx" % (self.base_dir, strftime)
        self.wb.save(save_path)
        return save_path


# print(ExcelUtil())
# if __name__ == '__main__':
#     excel_path = '%s/data/case_excel/接口测试用例.xlsx'%base_dir
#     result_path = '%s\\output\\run_result_excel'%base_dir + time.strftime("%Y%m%d_%H_%M_%S")
#     # ExcelUtil(excel_path).read_excel()
#     ExcelUtil(excel_path).write_excel()
